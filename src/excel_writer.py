"""
Writes one formatted report block to the master Excel workbook.

Behaviour:
  - Never deletes the master file — always appends below the last block
  - One sheet per calendar day, named {prefix}_{DD-Mon-YYYY}
  - At midnight a new sheet tab is created automatically
  - Each run appends exactly one block: separator + headers + data rows + TOTAL
  - KPI highlighting targets only the specific cell that triggered the alert:
      RED    = value below absolute floor threshold
      ORANGE = value dropped more than N% compared to previous row
  - All other cells keep normal alternating white/blue row colours
  - Column set is determined by the active PLMN via get_columns_for_plmn()
"""
import os
import logging
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from src.report_definitions import get_columns_for_plmn

log = logging.getLogger(__name__)


# ── Cell styling helpers ──────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _border() -> Border:
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _font(name="Arial", size=10, bold=False, color="000000") -> Font:
    return Font(name=name, size=size, bold=bold, color=color)

def _align(h="center", v="center", wrap=True) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _style(cell, font=None, fill=None, align=None, fmt=None):
    """Apply all style properties to a single cell."""
    if font:  cell.font          = font
    if fill:  cell.fill          = fill
    if align: cell.alignment     = align
    if fmt:   cell.number_format = fmt
    cell.border = _border()


# ── KPI flag detection ────────────────────────────────────────────────────────

def _compute_kpi_flags(summary, kpi_cols: list, thresholds: dict) -> dict:
    """
    Scans each KPI column and returns a flag for cells that breach thresholds.

    Returns dict of {(row_index, col_name): flag} where flag is:
      'drop'  → value dropped more than drop_pct% vs the previous row
      'floor' → value is below floor_pct absolute threshold
      'both'  → both conditions met (RED takes priority over ORANGE)

    Normalises \\n in column names so config.json keys always match
    even if the column name contains a line break character.
    """
    # Build normalised lookup so '\n' vs ' ' differences don't break matching
    norm_thr = {k.replace("\n", " "): v for k, v in thresholds.items()}

    flags = {}
    for col in kpi_cols:
        t = thresholds.get(col) or norm_thr.get(col.replace("\n", " "))
        if not t or col not in summary.columns:
            continue

        drop_pct  = float(t.get("drop_pct",  4.0))
        floor_pct = float(t.get("floor_pct", 40.0))
        vals      = summary[col].tolist()

        for i, val in enumerate(vals):
            if not isinstance(val, (int, float)):
                continue
            flag = None

            # Check absolute floor first
            if val < floor_pct:
                flag = "floor"

            # Check percentage drop vs previous row
            if i > 0:
                prev = vals[i - 1]
                if isinstance(prev, (int, float)) and prev > 0:
                    if (prev - val) / prev * 100 > drop_pct:
                        flag = "both" if flag == "floor" else "drop"

            if flag:
                flags[(i, col)] = flag

    return flags


# ── Sheet management ──────────────────────────────────────────────────────────

def _today_sheet_name(rdef: dict) -> str:
    """Generates the sheet tab name for today, e.g. 'Traffic_26-Feb-2026'."""
    return f"{rdef.get('sheet_prefix', 'Report')}_{datetime.now().strftime('%d-%b-%Y')}"[:31]


def _get_or_create_sheet(wb, rdef: dict, columns: list, widths: dict):
    """
    Returns an existing sheet for today or creates a new one.
    If the sheet already exists, positions the cursor 2 rows below
    the last used row (leaving one blank gap between blocks).
    """
    name = _today_sheet_name(rdef)
    if name in wb.sheetnames:
        ws       = wb[name]
        next_row = ws.max_row + 2   # one blank gap between report blocks
    else:
        ws = wb.create_sheet(title=name)
        ws.sheet_properties.tabColor = "1F3864"
        for ci, col in enumerate(columns, 1):
            ws.column_dimensions[get_column_letter(ci)].width = widths.get(col, 14)
        ws.freeze_panes = "B3"      # freeze date column and header row
        next_row = 1
        log.info(f"  Created new sheet: '{name}'")
    return ws, next_row


# ── Main writer ───────────────────────────────────────────────────────────────

def append_report(summary, rdef: dict, config: dict, plmn: str):
    """
    Appends one formatted report block to today's sheet in the master workbook.

    Block structure (top to bottom):
      Row 1:      Separator bar — report name, time window, generation timestamp
      Row 2:      Column headers — dark blue background, white text
      Rows 3–N:   Data rows — alternating white/blue, KPI cells highlighted
      Row N+1:    TOTAL row — SUM for counts, AVERAGE for rates

    Parameters:
      summary  : pd.DataFrame  from processor.process()
      rdef     : dict          report definition from report_definitions.py
      config   : dict          loaded from config.json
      plmn     : str           active PLMN (determines column set)

    Returns:
      (xlsx_path, sheet_name, first_row, last_row)
    """
    colors    = config["colors"]
    kpi_thr   = config["kpi_thresholds"]
    out_dir   = config["output"]["folder"]
    filename  = config["output"]["master_filename"]
    xlsx_path = os.path.join(out_dir, filename)
    os.makedirs(out_dir, exist_ok=True)

    # Get the correct column set for this PLMN
    columns   = get_columns_for_plmn(rdef, plmn)
    sum_cols  = set(rdef.get("sum_cols",  []))
    avg_cols  = set(rdef.get("avg_cols",  []))
    rate_cols = set(rdef.get("rate_cols", []))
    int_cols  = set(rdef.get("int_cols",  []))
    kpi_cols  = [c for c in rdef.get("kpi_cols", []) if c in columns]
    widths    = rdef.get("col_widths", {})
    flags     = _compute_kpi_flags(summary, kpi_cols, kpi_thr)

    # Open existing workbook or create a new one
    if os.path.exists(xlsx_path):
        wb = load_workbook(xlsx_path)
    else:
        wb = Workbook()
        wb.remove(wb.active)
        log.info(f"  Creating new workbook: {xlsx_path}")

    ws, next_row = _get_or_create_sheet(wb, rdef, columns, widths)

    # Prepare reusable style objects
    sep_font   = _font(size=11, bold=True,  color=colors["sep_font"])
    hdr_font   = _font(size=9,  bold=True,  color=colors["header_font"])
    data_font  = _font(size=9)
    total_font = _font(size=9,  bold=True)
    rate_font  = _font(size=9,  bold=True,  color="1F3864")

    sep_fill   = _fill(colors["sep_bg"])
    hdr_fill   = _fill(colors["header_bg"])
    total_fill = _fill(colors["total_row"])
    alt_fill   = _fill(colors["alt_row"])
    white_fill = _fill(colors["white"])
    drop_fill  = _fill(colors["drop_alert"])
    floor_fill = _fill(colors["floor_breach"])

    left_align   = _align("left",   wrap=False)
    center_align = _align("center")
    center_nw    = _align("center", wrap=False)

    # ── Separator / title row ─────────────────────────────────────────────────
    first_dt = summary["Date Time"].iloc[0]  if not summary.empty else "–"
    last_dt  = summary["Date Time"].iloc[-1] if not summary.empty else "–"
    now_str  = datetime.now().strftime("%d-%b-%Y %H:%M:%S")
    sep_label = (
        f"  {rdef['label']}  |  PLMN: {plmn}  |  "
        f"{first_dt}  →  {last_dt}  |  Generated: {now_str}"
    )

    sep_row = next_row
    ws.merge_cells(
        start_row=sep_row, start_column=1,
        end_row=sep_row,   end_column=len(columns)
    )
    c = ws.cell(row=sep_row, column=1, value=sep_label)
    _style(c, font=sep_font, fill=sep_fill, align=left_align)
    ws.row_dimensions[sep_row].height = 22

    # ── Column header row ─────────────────────────────────────────────────────
    hdr_row = sep_row + 1
    for ci, col in enumerate(columns, 1):
        c = ws.cell(row=hdr_row, column=ci, value=col)
        _style(c, font=hdr_font, fill=hdr_fill, align=center_align)
    ws.row_dimensions[hdr_row].height = 34

    # ── Data rows ─────────────────────────────────────────────────────────────
    data_start = hdr_row + 1
    for ri, row_data in summary.iterrows():
        excel_row = data_start + ri
        # Alternating row background — preserved regardless of KPI state
        row_fill  = alt_fill if ri % 2 == 0 else white_fill

        for ci, col in enumerate(columns, 1):
            val = row_data.get(col, "")
            c   = ws.cell(row=excel_row, column=ci, value=val)

            # Choose font and number format based on column type
            if col in rate_cols:
                fnt = rate_font;  fmt = "0.00"
            elif col in int_cols:
                fnt = data_font;  fmt = "#,##0"
            elif col == "Date Time":
                fnt = data_font;  fmt = "@"
            else:
                fnt = data_font;  fmt = "General"

            align = left_align if ci == 1 else center_nw

            # Only colour the exact cell that triggered a KPI flag.
            # All other cells use the normal alternating row background.
            flag = flags.get((ri, col))
            if flag in ("floor", "both"):
                fill = floor_fill   # RED
            elif flag == "drop":
                fill = drop_fill    # ORANGE
            else:
                fill = row_fill     # normal alternating

            _style(c, font=fnt, fill=fill, align=align, fmt=fmt)
        ws.row_dimensions[excel_row].height = 15

    # ── TOTAL row ─────────────────────────────────────────────────────────────
    total_row = data_start + len(summary)
    c = ws.cell(row=total_row, column=1, value="TOTAL")
    _style(c, font=total_font, fill=total_fill, align=left_align)

    for ci, col in enumerate(columns, 1):
        if ci == 1:
            continue
        c  = ws.cell(row=total_row, column=ci)
        cl = get_column_letter(ci)

        if col in sum_cols:
            # SUM of the data rows above
            c.value = f"=SUM({cl}{data_start}:{cl}{total_row - 1})"
            fmt     = "#,##0" if col in int_cols else "0.00"
        elif col in avg_cols:
            # AVERAGE of the data rows, ignoring errors
            c.value = f"=IFERROR(AVERAGE({cl}{data_start}:{cl}{total_row - 1}),0)"
            fmt     = "0.00"
        else:
            fmt = "General"

        _style(c, font=total_font, fill=total_fill, align=center_nw, fmt=fmt)
    ws.row_dimensions[total_row].height = 16

    wb.save(xlsx_path)
    sheet_name = _today_sheet_name(rdef)
    log.info(f"  Saved → sheet='{sheet_name}'  rows {sep_row}–{total_row}  |  {xlsx_path}")
    return xlsx_path, sheet_name, sep_row, total_row
